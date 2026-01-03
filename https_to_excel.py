#!/usr/bin/env python3
"""
HTTPS Data to Excel Writer
Fetches data from HTTPS endpoints and writes to Excel spreadsheets.
"""

import requests
import json
import ssl
import urllib3
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class HTTPSToExcel:
    """Fetch data from HTTPS endpoints and write to Excel files."""
    
    def __init__(self, verify_ssl=True, timeout=30):
        """
        Initialize the HTTPS to Excel writer.
        
        Args:
            verify_ssl: Whether to verify SSL certificates
            timeout: Request timeout in seconds
        """
        self.verify_ssl = verify_ssl
        self.timeout = timeout
        self.session = requests.Session()
        
        # Configure SSL
        if not verify_ssl:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    def fetch_json(self, url, headers=None, params=None):
        """
        Fetch JSON data from an HTTPS endpoint.
        
        Args:
            url: The HTTPS URL to fetch from
            headers: Optional headers dict
            params: Optional query parameters
            
        Returns:
            Parsed JSON data
        """
        try:
            response = self.session.get(
                url,
                headers=headers or {},
                params=params or {},
                verify=self.verify_ssl,
                timeout=self.timeout
            )
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Error fetching {url}: {e}")
            raise
    
    def fetch_csv(self, url, headers=None):
        """
        Fetch CSV data from an HTTPS endpoint.
        
        Args:
            url: The HTTPS URL to fetch from
            headers: Optional headers dict
            
        Returns:
            CSV content as string
        """
        try:
            response = self.session.get(
                url,
                headers=headers or {},
                verify=self.verify_ssl,
                timeout=self.timeout
            )
            response.raise_for_status()
            return response.text
        except requests.exceptions.RequestException as e:
            print(f"Error fetching {url}: {e}")
            raise
    
    def json_to_excel(self, data, output_path, sheet_name="Data"):
        """
        Write JSON data to Excel file.
        
        Args:
            data: JSON data (list of dicts or dict with nested data)
            output_path: Path for output Excel file
            sheet_name: Name of the worksheet
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Handle different JSON structures
        if isinstance(data, list) and len(data) > 0:
            # List of dictionaries
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
                self._write_headers(ws, headers)
                
                for row_idx, item in enumerate(data, start=2):
                    for col_idx, header in enumerate(headers, start=1):
                        value = item.get(header, "")
                        if isinstance(value, (dict, list)):
                            value = json.dumps(value)
                        ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                # Simple list
                for row_idx, item in enumerate(data, start=1):
                    ws.cell(row=row_idx, column=1, value=str(item))
                    
        elif isinstance(data, dict):
            # Single dictionary or nested structure
            if all(isinstance(v, list) for v in data.values()):
                # Multiple sheets for nested data
                for key, values in data.items():
                    if key == list(data.keys())[0]:
                        sheet = ws
                        sheet.title = key[:31]  # Excel sheet name limit
                    else:
                        sheet = wb.create_sheet(title=key[:31])
                    
                    if values and isinstance(values[0], dict):
                        headers = list(values[0].keys())
                        self._write_headers(sheet, headers)
                        
                        for row_idx, item in enumerate(values, start=2):
                            for col_idx, header in enumerate(headers, start=1):
                                value = item.get(header, "")
                                if isinstance(value, (dict, list)):
                                    value = json.dumps(value)
                                sheet.cell(row=row_idx, column=col_idx, value=value)
            else:
                # Simple key-value pairs
                self._write_headers(ws, ["Key", "Value"])
                for row_idx, (key, value) in enumerate(data.items(), start=2):
                    ws.cell(row=row_idx, column=1, value=str(key))
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value)
                    ws.cell(row=row_idx, column=2, value=str(value))
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Save workbook
        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")
        return output_path
    
    def _write_headers(self, ws, headers):
        """Write styled headers to worksheet."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def fetch_and_write(self, url, output_path, sheet_name="Data", headers=None, params=None):
        """
        Fetch data from HTTPS and write to Excel in one step.
        
        Args:
            url: HTTPS URL to fetch from
            output_path: Path for output Excel file
            sheet_name: Name of the worksheet
            headers: Optional request headers
            params: Optional query parameters
            
        Returns:
            Path to created Excel file
        """
        data = self.fetch_json(url, headers=headers, params=params)
        return self.json_to_excel(data, output_path, sheet_name)
    
    def fetch_multiple_to_excel(self, urls_config, output_path):
        """
        Fetch from multiple HTTPS endpoints and combine into one Excel file.
        
        Args:
            urls_config: List of dicts with 'url', 'sheet_name', 'headers', 'params'
            output_path: Path for output Excel file
            
        Returns:
            Path to created Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        
        wb = Workbook()
        first_sheet = True
        
        for config in urls_config:
            url = config['url']
            sheet_name = config.get('sheet_name', 'Data')[:31]
            headers = config.get('headers')
            params = config.get('params')
            
            try:
                data = self.fetch_json(url, headers=headers, params=params)
                
                if first_sheet:
                    ws = wb.active
                    ws.title = sheet_name
                    first_sheet = False
                else:
                    ws = wb.create_sheet(title=sheet_name)
                
                # Write data to sheet
                if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                    sheet_headers = list(data[0].keys())
                    self._write_headers(ws, sheet_headers)
                    
                    for row_idx, item in enumerate(data, start=2):
                        for col_idx, header in enumerate(sheet_headers, start=1):
                            value = item.get(header, "")
                            if isinstance(value, (dict, list)):
                                value = json.dumps(value)
                            ws.cell(row=row_idx, column=col_idx, value=value)
                
                self._auto_adjust_columns(ws)
                print(f"Added sheet '{sheet_name}' from {url}")
                
            except Exception as e:
                print(f"Error processing {url}: {e}")
                continue
        
        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")
        return output_path


def main():
    """Example usage of HTTPSToExcel."""
    
    # Initialize the writer
    writer = HTTPSToExcel(verify_ssl=True, timeout=30)
    
    # Example 1: Fetch from a single API endpoint
    print("=" * 60)
    print("HTTPS to Excel Writer - Examples")
    print("=" * 60)
    
    # Example with JSONPlaceholder API (public test API)
    try:
        print("\n[Example 1] Fetching users from JSONPlaceholder API...")
        writer.fetch_and_write(
            url="https://jsonplaceholder.typicode.com/users",
            output_path="users_data.xlsx",
            sheet_name="Users"
        )
    except Exception as e:
        print(f"Example 1 failed: {e}")
    
    # Example 2: Fetch posts
    try:
        print("\n[Example 2] Fetching posts...")
        writer.fetch_and_write(
            url="https://jsonplaceholder.typicode.com/posts",
            output_path="posts_data.xlsx",
            sheet_name="Posts"
        )
    except Exception as e:
        print(f"Example 2 failed: {e}")
    
    # Example 3: Multiple endpoints to single Excel file
    try:
        print("\n[Example 3] Fetching multiple endpoints to one file...")
        urls_config = [
            {
                'url': 'https://jsonplaceholder.typicode.com/users',
                'sheet_name': 'Users'
            },
            {
                'url': 'https://jsonplaceholder.typicode.com/posts',
                'sheet_name': 'Posts'
            },
            {
                'url': 'https://jsonplaceholder.typicode.com/comments',
                'sheet_name': 'Comments'
            }
        ]
        writer.fetch_multiple_to_excel(urls_config, "combined_data.xlsx")
    except Exception as e:
        print(f"Example 3 failed: {e}")
    
    # Example 4: Custom data to Excel
    print("\n[Example 4] Writing custom data to Excel...")
    custom_data = [
        {"ID": 1, "Name": "Server A", "Status": "Online", "CPU": "45%", "Memory": "62%"},
        {"ID": 2, "Name": "Server B", "Status": "Online", "CPU": "32%", "Memory": "48%"},
        {"ID": 3, "Name": "Server C", "Status": "Offline", "CPU": "0%", "Memory": "0%"},
        {"ID": 4, "Name": "Server D", "Status": "Online", "CPU": "78%", "Memory": "85%"},
    ]
    writer.json_to_excel(custom_data, "server_status.xlsx", sheet_name="Servers")
    
    print("\n" + "=" * 60)
    print("All examples completed!")
    print("=" * 60)


if __name__ == "__main__":
    main()
